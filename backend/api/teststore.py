import sys
import shutil
from django.http import HttpResponse
from .serializers import *
from .models import *
import random, os
import numpy as np
from . import json
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist
from rest_framework.response import Response
from django.core import serializers
from rest_framework import generics
from django.views.generic import ListView
import re
import pandas as pd
import pdb
import xlrd
import xgboost as xgb
import statsmodels.api as sm
from scipy.stats import ttest_rel,ttest_ind

from sklearn.metrics import r2_score
from sklearn.feature_selection import chi2
from scipy.stats import ttest_ind, norm
from statsmodels.stats.power import TTestIndPower
from sklearn.preprocessing import StandardScaler, normalize

from .gower_function import gower_distances as gf
from .helpers import mkdir, chmod, getFeatures, reduce, logs, datetime, init
from .helpers import  storeChar, config, storeKeys, constrainCols, analyseFeatureCols, identifyControlCols, liftAnalysisCols, logIndex,trendCols

from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import string
import xlsxwriter
from openpyxl import load_workbook
import shutil

rootDir = os.path.dirname(os.path.abspath(__file__))
np.set_printoptions(suppress=True)

random.seed(999)
dataStore = None
success = "Success"

init()


def ANALYSEHELPER():
	"""
        params:
        ------
            startDate       :   start date for feature selection - 'yyyy-mm-dd'                 -   string
            endDate         :   end date for feature selection - 'yyyy-mm-dd'                   -   string

        returns:
        -------
            dataStore   :   store x week data for all stores filtered for the selected interval -   pandas DataFrame

        dependencies:
        ------------
            dataStore.csv   :   file containing store x week data located at the path           -   .csv file
                                mentioned in config.ini as ['dataPath']
    """
	global dataStore
	try:
		dataStore = pd.read_csv(os.path.join(config['pathConfig']['dataPath'], config['fileConfig']['dataStore']))
		dataStore[storeKeys] = dataStore[storeKeys].astype('int').astype('str')
		dataStore['wkEndDate'] = pd.to_datetime(dataStore['wkEndDate'])
		dataStore = dataStore.sort_values(by=['wkEndDate'])
	except:
		if dataStore is None:
			exit()
		else:
			# TODO: send "update fail status" to user
			pass
	return dataStore



if dataStore is None:
	dataStore = ANALYSEHELPER()


def dateRange(request):
	results =  {
		"startdate": str(dataStore.wkEndDate.unique()[0]),
		"enddate"  : str(dataStore.wkEndDate.unique()[-1])
	}
	return json.Response(results,True)

	# minDate = pd.to_datetime(dataStore.wkEndDate).drop_duplicates().min()
	# maxDate = pd.to_datetime(dataStore.wkEndDate).drop_duplicates().max()

def Weekduration(request):
	import json as j
	data = j.loads(request.body)
	try:
		start = data['start']
		end = data['end']
		start = str(start['year']) + "-" + str(start['month']) + "-" + str(start['day'])
		end = str(end['year']) + "-" + str(end['month']) + "-" + str(end['day'])
		start = startEnd(start)
		end = startEnd(end, end=True)
		result = {
			"startdate" : str(start),
			"enddate"   : str(end),
			"interval"  : str(int(((end - start).days + 1) / 7))
		}
		# print(start, end, int(((end - start).days + 1) / 7))
		return  json.Response(result,True)
	except:
		return json.Response("Can't get Duration in weeks", False)


def startEnd(date, end=False):
	import datetime
	day = 0
	if end:
		day = 6
	date = datetime.datetime.strptime(date, '%Y-%m-%d')
	return date - datetime.timedelta(days=(date.weekday() - day) % 7)


class TestStore(ListView):
	
	# Get All TestStore Data

	def post(self,request):
		import json as j 
		data = j.loads(request.body)
		data = data['data']
		get_storedata = StoreMstr.objects.filter(is_active=True,is_deleted=False).order_by('store_sk')
		if(data['state']!=""):
			if(len(data['state'])>0):
				get_storedata = get_storedata.filter(state_long__in=data['state'])
		if(data['storetype']!=""):
			if(len(data['storetype'])>0):
				get_storedata = get_storedata.filter(store_type__in=data['storetype'])
		store_data = GetAllStoreSerializer(get_storedata,many=True)	
		return json.Response(store_data.data,True)	
	
def GetAllSavedData(request):

	import json as j 
	data = j.loads(request.body)
	data = data['data']
	gettestStore = TestMstr.objects.get(test_name=data,is_active=True,is_deleted=False)
	results  = gettestStoreSerializer(gettestStore)
	return json.Response(results.data,True)

def Get_StoresDetails(request):

	import json as j 
	data = j.loads(request.body)
	data = data['data']
	gettestStore = StoreMstr.objects.filter(store_sk__in=data)
	results  = MatchTestStoreSerializer(gettestStore,many=True)
	return json.Response(results.data,True)

def FromMeasurement(request):

	import json as j 
	data = j.loads(request.body)
	datas = data['data']
	
	market_id = datas['market_id']
	test_name = datas['test_name']
	check_name = TestMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,test_name=test_name).count()
	if(check_name >0):
		return json.Response('Test Name Already exists',False)
	else:

		Market_id = MarketMstr.objects.get(is_active=True,is_deleted=False,market_id=datas['market_id'])
		StageSave = TestMstr(
			test_name   = datas['test_name'],
			market_id   = Market_id,
			stage_id    = datas['stage_id'],
			created_on	= int(time.time()),
			is_deleted 	= False,
			is_active	= True
			)
		StageSave.save()

		Record_save = RecordMstr(
			test_id   = StageSave.test_id,
			stage_id  = datas['stage_id'],
			record_value = datas['stringified_data'],
			market_id  = Market_id,
			created_on	= int(time.time()),
			is_deleted 	= False,
			is_active	= True
			)
		Record_save.save()

		return json.Response(StageSave.pk,True)

	# gettestStore = StoreMstr.objects.filter(store_sk__in=data)
	# results  = MatchTestStoreSerializer(gettestStore,many=True)
	# return json.Response(results.data,True)

class Edit_TestStore(ListView):

	def get(self,serializer,pk):	
		gettestStore = TestMstr.objects.get(test_id=pk)
		results  = gettestStoreSerializer(gettestStore)
		return json.Response(results.data,True)
		# try:
		# 	gettestStore = TestMstr.objects.get(test_id=pk)
		# 	results  = gettestStoreSerializer(gettestStore)
			
		# except:
		# 	return json.Response('Can`t able to get data',False)


	def delete(self,serializer,pk):	
		try:
			deletetestStore = TestMstr.objects.filter(test_id=pk).update(deleted_at=int(time.time()),is_deleted=True,is_active=False)
			return json.Response('Test Deleted Successfully',True)
		except:
			return json.Response('Can`t able to delete data',False)

# Check Test Name with Existing Table

def Checkname(request):
	import json as j 
	data = j.loads(request.body)
	market_id = data['market_id']
	test_name = data['test_name']
	check_name = TestMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,test_name=test_name).count()
	if(check_name >0):
		return json.Response('Test Name Already exists',False)
	else:
		return json.Response('Test Name Available',True)

## Get All Category from DB

def Get_category(request):
	import json as j 
	data = j.loads(request.body)
	market_id = data['market_id']
	hierarchy_id = data['hierarchy_id']	
	if(hierarchy_id==1):
		getcategory = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True).values('levelone_cate').distinct()
		Category_data = GetLeveloneSerializer(getcategory,many=True)	
	elif(hierarchy_id==2):
		getcategory = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True).values('leveltwo_cate').distinct()
		Category_data = GetLeveltwoSerializer(getcategory,many=True)	
	elif(hierarchy_id==3):
		getcategory = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True).values('levelthree_cate').distinct()
		Category_data = GetLevelthreeSerializer(getcategory,many=True)	
	elif(hierarchy_id==4):
		getcategory = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True).values('levelfour_cate').distinct()
		Category_data = GetLevelfourSerializer(getcategory,many=True)	
	else:
		return json.Response('Invalid Hierarchy Id',False)

	try:	
		return json.Response(Category_data.data,True)
	except:
		return json.Response('Could`t find Category Name',False)


## Get All Products based on Categories from DB

def Get_products(request):
	import json as j 
	data = j.loads(request.body)
	market_id = data['market_id']
	hierarchy_id = data['hierarchy_id']	
	category_desc = data['category_desc']	

	if(hierarchy_id==1):
		getproducts = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,levelone_cate__in=category_desc)		
	elif(hierarchy_id==2):
		getproducts = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,leveltwo_cate__in=category_desc)
	elif(hierarchy_id==3):
		getproducts = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,levelthree_cate__in=category_desc)
	elif(hierarchy_id==4):
		getproducts = ProductMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,levelfour_cate__in=category_desc)
	else:
		return json.Response('Invalid Hierarchy Id',False)

	product_item = GetproductSerializer(getproducts,many=True)
	try:	
		return json.Response(product_item.data,True)
	except:
		return json.Response('Could`t find Products Name',False)


## Get All Features from DB


def Getall_features(request):
	import json as j 
	data = j.loads(request.body)
	market_id = data['market_id']	
	getallfeatures = FeatureMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True).order_by('feat_name')			
	features = GetAllFeaturesSerializer(getallfeatures,many=True)
	try:	
		return json.Response(features.data,True)
	except:
		return json.Response('Could`t find Features Name',False)




def randomString(stringLength=7):
    """Generate a random string of fixed length """
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))

# Get matching Test Store with Upload Excel

def GetMatchingTestStore(request):
	
	filename = request.FILES['match_store']
	temp_filename =""
	if filename:
		filename_check = filename.name
		ext = [".xlsx", ".csv", ".xls"]
		if filename_check.endswith(tuple(ext)):
			extesnion = os.path.splitext(filename_check)[1]
			if(extesnion=='.csv'):
				data = pd.read_csv(filename)
			else:
				data = pd.read_excel(filename)
			MatchStore_id = []
			UnMatch_id = []
			for i in data.columns:
				if(data.columns[0]!="Test Store ID (store_sk)"):
					return json.Response("Invalid file input! Please upload again",False)
			for i in data.index:
				if(data['Test Store ID (store_sk)'][i]==None):
					break;
				else:
					try:
						check_store = StoreMstr.objects.filter(store_sk=data['Test Store ID (store_sk)'][i],is_deleted=False,is_active=True).count()
						if(check_store>0):
							MatchStore_id.append(data['Test Store ID (store_sk)'][i])
						else:
							UnMatch_id.append(data['Test Store ID (store_sk)'][i])
					except ValueError:
						UnMatch_id.append(data['Test Store ID (store_sk)'][i])

			temp_filename = randomString() + ".xlsx"
			path = default_storage.save(temp_filename, ContentFile(filename.read()))
			tmp_file = os.path.join(settings.TESTSTORES_DIR, path)

			GetMatchStore = StoreMstr.objects.filter(store_sk__in=MatchStore_id,is_active=True,is_deleted=False).order_by('store_sk')
			Store_data  = MatchTestStoreSerializer(GetMatchStore,many=True)
			results = {
					'match_data'   : Store_data.data,
					'unmatch_data' : str(UnMatch_id),
					'temp_filename': temp_filename
			} 
			return json.Response(results,True)
		else:
			return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
	else:
		return json.Response("File Upload Failed",False)


def UploadTestControlStore(request):
	
	filename = request.FILES['testcontrol_store']
	testname = request.POST['testname']
	market_id = request.POST['market_id']
	check_name = TestMstr.objects.filter(market_id=market_id,is_deleted=False,is_active=True,test_name=testname).count()
	if(check_name >0):
		return json.Response('Test Name Already exists',False)
	
	if filename:
		filename_check = filename.name
		ext = [".xlsx", ".csv", ".xls"]
		if filename_check.endswith(tuple(ext)):
			extesnion = os.path.splitext(filename_check)[1]
			if (extesnion == '.csv'):
				data = pd.read_csv(filename)
			else:
				data = pd.read_excel(filename)
			temp_filename = randomString()+".xlsx"
			path = default_storage.save(temp_filename, ContentFile(filename.read()))
			tmp_file = os.path.join(settings.TESTCONTROLSTORES_DIR, path)
			MatchStore_id = []
			UnMatch_id = []
			MatchStore_id1 = []
			UnMatch_id1 = []
			
			if(data.columns[0]!="Test store ID (store_sk)"):
				return json.Response("Invalid file input! Please upload again",False)
			if (data.columns[1] != "Control store ID (store_sk)"):
				return json.Response("Invalid file input! Please upload again", False)
			if (data.columns[2] != "Rank"):
				return json.Response("Rank Column missing! Please upload again", False)
			
			testStores =[]
			for i in data.index:
				if(data['Test store ID (store_sk)'][i]==None):
					break;
				else:
					check_store = StoreMstr.objects.filter(store_sk=data['Test store ID (store_sk)'][i],is_deleted=False,is_active=True).count()
					if(check_store>0):
						GetControlMatchStore = StoreMstr.objects.get(store_sk=data['Test store ID (store_sk)'][i],is_active=True,is_deleted=False)
						Store_data1  = MatchTestStoreSerializer(GetControlMatchStore).data
						check_store1 = StoreMstr.objects.filter(store_sk=data['Control store ID (store_sk)'][i],is_deleted=False,is_active=True).count()
						if(check_store1>0):
							GetControlMatchStore1 = StoreMstr.objects.get(store_sk=data['Control store ID (store_sk)'][i],is_active=True,is_deleted=False)
							Store_data2  = MatchTestStoreSerializer(GetControlMatchStore1).data
							testStores.append({"teststore_id":Store_data1['store_sk'],"teststore_name":Store_data1['store_name'],"controlstore_id":Store_data2['store_sk'],"controlstore_name":Store_data2['store_name'],"rank" : int(data['Rank'][i])})

			finaldata = data.reset_index().to_json(orient='records')
			restuls ={
			"stores" : testStores,
			"data" : finaldata
			}
			return json.Response(restuls,True)
		else:
			return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
	else:
		return json.Response("File Upload Failed",False)

	
# Exclude Control Store with Upload Excel

def ExcludeControlStore(request):

	filename = request.FILES['exclude_store']
	teststore = request.POST['teststore']
	teststore  = teststore.split(",")

	if filename:
		filename_check = filename.name
		ext = [".xlsx", ".csv", ".xls"]
		if filename_check.endswith(tuple(ext)):
			extesnion = os.path.splitext(filename_check)[1]
			if (extesnion == '.csv'):
				data = pd.read_csv(filename)
			else:
				data = pd.read_excel(filename)
			MatchStore_id = []
			UnMatch_id = []
			for i in data.columns:
				if(data.columns[0]!="control store id (store_sk)"):
					return json.Response("Invalid file input! Please upload again",False)
			for i in data.index:
				if(data['control store id (store_sk)'][i]==None):
					break;
				else:
					try:
						check_store = StoreMstr.objects.filter(store_sk=data['control store id (store_sk)'][i],is_deleted=False,is_active=True).count()
						if(check_store>0):
							if (data['control store id (store_sk)'][i] not in MatchStore_id):
								MatchStore_id.append(data['control store id (store_sk)'][i])
						else:
							if (data['control store id (store_sk)'][i] not in UnMatch_id):
								UnMatch_id.append(data['control store id (store_sk)'][i])
					except ValueError:
						UnMatch_id.append(data['control store id (store_sk)'][i])
			GetMatchStore = StoreMstr.objects.filter(store_sk__in=MatchStore_id,is_active=True,is_deleted=False).order_by('store_sk')
			Store_data  = MatchTestStoreSerializer(GetMatchStore,many=True)
			results = {
					'match_data'   : Store_data.data,
					'unmatch_data' : str(UnMatch_id)
			} 
			return json.Response(results,True)
		else:
			return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
	else:
		return json.Response("File Upload Failed",False)


# Upload Excel Test Stores
def UploadTestStore(request):
		filename = request.FILES['store_file']
		market   = request.POST['market_id']
		if filename:
			filename_check = filename.name
			ext = [".xlsx", ".csv", ".xls"]
			if filename_check.endswith(tuple(ext)):
				extesnion = os.path.splitext(filename_check)[1]
				if (extesnion == '.csv'):
					data = pd.read_csv(filename)
				else:
					data = pd.read_excel(filename)
				# try:
				Market_id = MarketMstr.objects.get(is_active=True,is_deleted=False,market_id=market)

				for i in data.index:
					if(data['storeID'][i]==None):
						break;
					else:
						Save_TestStore = StoreMstr(
							market_id 						= Market_id,
							store_name                      = data['storeName'][i],
							store_name_short                = data['storeName'][i],  
							store_sk                        = data['storeID'][i],
							state_short                     = data['state'][i],
							state_long                      = data['state'][i],
							store_type                      = data['store_type'][i],
							created_on						= int(time.time()),
							is_deleted 						= False,
							is_active						= True
							)
						Save_TestStore.save()

				return json.Response('Test Store Updated Successfully',True)	
				# except:
					# return json.Response('Can`t able to upload', False) 
			else:
				return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
		else:
			return json.Response("File Upload Failed",False)

# Upload Products with Category Mapping Test Stores
def UploadItemMapping(request):
		filename = request.FILES['menu_items']
		data = pd.read_excel(filename)
		market = request.POST['market_id']
		Market_id = MarketMstr.objects.get(is_active=True,is_deleted=False,market_id=market)
		try:
			for i in data.index:
				if(data['itemID'][i]==None):
					break;
				else:				
					Save_items = ProductMstr(
					market_id 						= Market_id,
					menu_item_bk                   	= data['itemID'][i],
					menu_item_desc                 	= data['itemDes'][i],
					# size_code                       = data['size_code'][i],
					# levelone_cate                  	= data['hierarchy1'][i],
					leveltwo_cate                  	= data['hierarchy2'][i],
					# levelthree_cate                	= data['hierarchy3'][i],
					# levelfour_cate                	= data['hierarchy4'][i],
					)
		
				Save_items.save()

			return json.Response('Menu Items Uploaded Successfully',True)	
		except:
			return json.Response('Can`t able to upload', False) 




# Check Additional Feature for store sk and feature percentage

def Check_additionalFeature(request):
	filename = request.FILES['check_feature']
	if filename:
		filename_check = filename.name
		ext = [".xlsx", ".csv", ".xls"]
		if filename_check.endswith(tuple(ext)):
			extesnion = os.path.splitext(filename_check)[1]
			if (extesnion == '.csv'):
				data = pd.read_csv(filename)
			else:
				data = pd.read_excel(filename)
			MatchStore_id = []
			UnMatch_id = []
			feature_name =[]
			unmatchted_feature =[]
			for i in data.columns:
				if(i=="store_sk"):
					pass
				else:
					feature_name.append(i)
				if(data.columns[0]!="store_sk"):
					return json.Response("Invalid file input! Please upload again",False)
			for i in data.index:
				if(data['store_sk'][i]==None):
					break;
				else:
					try:
						checkall_store = StoreMstr.objects.filter(is_deleted=False,is_active=True).count()
						check_store = StoreMstr.objects.filter(store_sk=data['store_sk'][i],is_deleted=False,is_active=True).count()
						if(check_store>0):
							MatchStore_id.append(data['store_sk'][i])
						else:
							UnMatch_id.append(data['store_sk'][i])
					except ValueError:
						UnMatch_id.append(data['store_sk'][i])
			for val in feature_name:
				checkfeature = FeatureTbl.objects.filter(feature_name = val,is_deleted=False, is_active=True).count()
				if(checkfeature>0):
					unmatchted_feature.append(val)

			results = {
				"AllTest_store" : checkall_store,
				"Match_store"   : len(MatchStore_id),
				"UnMatch_id"    : len(UnMatch_id),
				"Additional_Features"  : feature_name,
				"Match_features" : unmatchted_feature,
				"Additional_Feature_data" :  data.to_json(orient='records')
			}
			return json.Response(results,True)
		else:
			return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
	else:
		return json.Response("File Upload Failed",False)
	





# Upload Considered Features Master and Slave Tables

# def UploadFeatures(request):
# 	filename = request.FILES['feature_data']
# 	if filename:
# 		filename_check = filename.name
# 		ext = [".xlsx", ".csv", ".xls"]
# 		if filename_check.endswith(tuple(ext)):
# 			extesnion = os.path.splitext(filename_check)[1]
# 			if (extesnion == '.csv'):
# 				data = pd.read_csv(filename)
# 			else:
# 				data = pd.read_excel(filename)
# 			market = request.POST['market_id']
# 			Market_id = MarketMstr.objects.get(is_active=True,is_deleted=False,market_id=market)
# 			try:

# 				# Remove Old Feature Mst Data
# 				count_feature = FeatureMstr.objects.filter(is_active=True,is_deleted=False,market_id=market).count()
# 				if(count_feature>0):
# 					remove_feature = FeatureMstr.objects.get(is_active=True,is_deleted=False,market_id=market).delete()
# 				for i in data.columns:
# 					FeatureMst = FeatureMstr(
# 						feat_name   = i,
# 						market_id   = Market_id,
# 						created_on	= int(time.time()),
# 						is_deleted 	= False,
# 						is_active	= True
# 						)
# 					FeatureMst.save()	
# 				return json.Response('Feature Items Uploaded Successfully',True)	
# 			except:
# 				return json.Response('Can`t able to upload', False)
# 		else:
# 			return json.Response("Invalid File Format! Please upload .csv or .xlsx",False)
# 	else:
# 		return json.Response("File Upload Failed",False)






def UploadFeatures(request):
	filename = request.FILES['feature_data']
	data = pd.read_excel(filename)
	feat_id = FeatureMstr.objects.get(is_active=True,is_deleted=False,feat_id=4)
	for i in data.index:
		FeatureMst = FeatureTbl(
			feature_name   = data['feature_name'][i],
			feat_id   	= feat_id,
			created_on	= int(time.time()),
			is_deleted 	= False,
			is_active	= True
			)
		FeatureMst.save()	
	return True


def SaveStage(request):
	import json as j 
	data = j.loads(request.body)
	datas = data['data']
	try:		
		Market_id = MarketMstr.objects.get(is_active=True,is_deleted=False,market_id=datas['market_id'])
		StageSave = TestMstr(
			test_name   = datas['test_name'],
			market_id   = Market_id,
			stage_id    = datas['stage_id'],
			created_on	= int(time.time()),
			is_deleted 	= False,
			is_active	= True
			)
		StageSave.save()

		remove_record = RecordMstr.objects.filter(test_id = StageSave.test_id).delete()

		Record_save = RecordMstr(
			test_id   = StageSave.test_id,
			stage_id  = datas['stage_id'],
			record_value = datas['stringified_data'],
			market_id  = Market_id,
			created_on	= int(time.time()),
			is_deleted 	= False,
			is_active	= True
			)
		Record_save.save()

		return json.Response(StageSave.pk,True)
	except:
		return json.Response('Unable to Save',False)



def UpdateStage(request):
	import json as j
	data = j.loads(request.body)
	datas = data['data']
	stringify = data['stringified_data']
	# Market_id = TestMstr.objects.filter(is_active=True, is_deleted=False, test_id=datas['trial']).update(stage_id=datas['stage_id'])
	# return json.Response('Stage saved Successfully', True)
	# try:
	removee_rec = RecordMstr.objects.filter(is_active=True, is_deleted=False, test_id=datas['trial'],stage_id=datas['stage_id']).count()
	if(removee_rec>0):
		removee_rec = RecordMstr.objects.filter(is_active=True, is_deleted=False, test_id=datas['trial'],
												stage_id=datas['stage_id']).delete()

	StageSave = RecordMstr(
		test_id=datas['trial'],
		stage_id=datas['stage_id'],
		record_value= stringify,
		created_on=int(time.time()),
		is_deleted = False,
		is_active	= True
	)
	StageSave.save()

	UpdateStage = TestMstr.objects.filter(is_active=True, is_deleted=False, test_id=datas['trial']).update(stage_id=datas['stage_id'])

	return json.Response(StageSave.pk ,True)
	# except:
	# 	return json.Response('Unable to Save' ,False)

			# deletetestStore = TestMstr.objects.filter(test_id=pk).update(deleted_at=int(time.time()),is_deleted=True,is_active=False)
			# return json.Response('Test Deleted Successfully',True)
		# except:
		# 	return json.Response('Can`t able to delete data',False)



def LoadSaveData(request):	
	GetloadData = TestMstr.objects.filter(is_active=True,is_deleted=False).order_by('-created_on','-modified_on')
	results = LoadDataSeralizer(GetloadData,many=True)
	try:
		return json.Response(results.data,True)
	except:
		return json.Response('Can`t able to get data',False)

	
### Tiger Code


class Downloadreport(ListView):

	def get(self,serializer,pk):
		import json as j
		gettestStore = TestMstr.objects.get(test_id=pk)
		results = gettestStoreSerializer(gettestStore)
		finaldata = results.data

		first_record = dict(finaldata['records'][0])
		fir = first_record['record_value']
		stores_json = j.loads(fir)
		teststores = stores_json['select_store']

		second_record = dict(finaldata['records'][1])
		second = second_record['record_value']
		load_jsons = j.loads(second)

		third_record = dict(finaldata['records'][2])
		third = third_record['record_value']
		load_jsons1 = j.loads(third)

		analy_str = load_jsons['analystring']
		categor_features = analy_str['categor_features']

		analy_str = load_jsons['analystring']
		contiue_features= analy_str['contiue_features']

		report_name = "Trial_" + stores_json['test_name'] + "_Summary_Report.xlsx"

		workbook = xlsxwriter.Workbook(report_name)
		checkall_store = StoreMstr.objects.filter(is_deleted=False,is_active=True).count()


		# First Sheet Summary
		worksheet = workbook.add_worksheet('Summary')
		row = 0
		col = 0
		market = stores_json['market_id']
		target_metric = load_jsons['target_metric']

		if(market==1):
			market_name = "Australia"
		else:
			pass
		worksheet.write(row, col, "Market -" + market_name)
		worksheet.write(row+2, col, "Trial -" +  stores_json['test_name'] + " Summary")
		worksheet.write(row+4, col, "Total " + str(checkall_store) + " Store Considered for analysis. " + str(len(teststores)) + " test stores")
		worksheet.write(row+6, col, "Data table used to build pipeline")
		worksheet.write(row+7, col, "Store Charachteristics")
		worksheet.write(row+8, col, "Demographics")
		worksheet.write(row+9, col, "Weekly Pmix")
		worksheet.write(row+10, col, "Weekly sales and gcs")
		worksheet.write(row+11, col, "Weekly financials")
		worksheet.write(row+ 13, col, "Store excluded due to insufficient data ") # to do exclude test stores
		if(target_metric=='1'):
			target_name = "Wt Avg price"
		else:
			target_name = "% Sales"
		worksheet.write(row + 13, col, "Target Metric for Feature Identification is "+target_name +" for " + str(len(load_jsons['prd'])) +" Products listed under " +
						str(len(load_jsons['prd_cat'])) +" Product categories in Hierachy Level " + str(load_jsons['Hierarchy']))

		worksheet.write(row + 15, col, "Feature Selection Window - Start Date " + str(load_jsons['Startdt']) + "/ End Date " + str(load_jsons['Enddt']) + " Duration of the window in weeks - " +
						str(load_jsons['duration_window'])) #

		worksheet.write(row + 17, col, "List of Considered Continuous " + str(len(categor_features)) + " and Categorical " + str(len(contiue_features)))

		worksheet.write(row + 19, col,"List of Additional Features user input ") # to do additional features

		worksheet.write(row + 21, col, "Selected Features Top 7 Features - by Descending information Gain with values")
		worksheet.write(row + 21, col, "Selected Features")
		worksheet.write(row + 21, col+1, "Information Gain")
		worksheet.write(row + 21, col+2, "Forced Variable Flag")

		row = 22
		col = 0
		for stores in load_jsons['selected_feat']:
			worksheet.write(row, col, stores)
			row += 1

		worksheet.write(row+2 , col, "Model Adjusted R-Squared " + load_jsons['adjusted_r_squared'])

		worksheet.write(row +4, col, "Match "+ str(load_jsons['no_of_cntrl']) + " control stores for each test stores")

		worksheet.write(row + 6, col, "Control Stores to exclude ") ### to do exclude test stores

		worksheet.write(row + 8, col, "Matched Results - Hypothesis testing")

		worksheet.write(row + 9, col, "Higher p-values indicate tes(t stores and matched control stores are similar based on the top features selected.")

		worksheet.write(row + 12, col, "Variable Name")

		col=1
		for cou in range(int(load_jsons['no_of_cntrl'])):
			worksheet.write(row+12, col, "Control -" + str(col) + " p-value (avg p-value of control -" + str(col) +" stores)")
			col += 1

		variable_name = load_jsons['stat_results']

		row = row +11
		col = 0
		for value in variable_name:
			worksheet.write(row, col, value['VariableName'])
			if(load_jsons['no_of_cntrl']=="5"):
				worksheet.write(row, col+1, value['control_1'])
				worksheet.write(row, col+2, value['control_2'])
				worksheet.write(row, col+3, value['control_3'])
				worksheet.write(row, col+4, value['control_4'])
				worksheet.write(row, col+5, value['control_5'])
			if(load_jsons['no_of_cntrl']=="4"):
				worksheet.write(row, col + 1, value['control_1'])
				worksheet.write(row, col + 2, value['control_2'])
				worksheet.write(row, col + 3, value['control_3'])
				worksheet.write(row, col + 4, value['control_4'])
			if (load_jsons['no_of_cntrl'] == "3"):
				worksheet.write(row, col + 1, value['control_1'])
				worksheet.write(row, col + 2, value['control_2'])
				worksheet.write(row, col + 3, value['control_3'])
			if (load_jsons['no_of_cntrl'] == "2"):
				worksheet.write(row, col + 1, value['control_1'])
				worksheet.write(row, col + 2, value['control_2'])
			if (load_jsons['no_of_cntrl'] == "1"):
				worksheet.write(row, col + 1, value['control_1'])
			row +=1

		worksheet.write(row+1, col, "Test Measurement - Hypothesis Testing")
		worksheet.write(row +2 , col, "Lower p-values indicate selected test stores perform significantly better than the matched control stores based on measurement metrics selected")

		worksheet.write(row + 5, col,"Measurement metric")
		worksheet.write(row + 5, col+1, "P -value")
		worksheet.write(row + 5, col+2, "Lift")

		row = row + 6


		for metri in load_jsons1['variable']:
			worksheet.write(row, col, metri)
			worksheet.write(row, col+1, load_jsons1['metric'][col])
			worksheet.write(row, col+2, load_jsons1['delta'][col])
			row += 1

		# Second Sheet Test Store IDS

		worksheet = workbook.add_worksheet('Test Store List')
		row = 0
		col = 0

		for stores in teststores:
			worksheet.write(row, col, stores)
			row += 1
		###  End of Second Sheet

		## Third Sheet Product and  Category

		prd_cat = load_jsons['prd_cat']
		prd = load_jsons['prd']
		worksheet = workbook.add_worksheet('Product Category - Products')
		row = 1
		col = 0

		worksheet.write(0, 0, 'Test Products Category')

		for category in prd_cat:
			worksheet.write(row, col, category)
			# worksheet.write(row, col + 1, elm2)
			row += 1

		worksheet.write(0, 1, 'Test Products')
		row = 1
		col1 = 1
		for product in prd:
			worksheet.write(row, col1, product)
			# worksheet.write(row, col + 1, elm2)
			row += 1

		## Fourth Sheet continues Variable

		worksheet = workbook.add_worksheet('Continues Variable Considered')
		row = 1
		col = 0

		worksheet.write(0, 0, 'Continous Variables considered for feature selection')

		for category in contiue_features:
			worksheet.write(row, col, category)
			# worksheet.write(row, col + 1, elm2)
			row += 1

		## Fifth Sheet Categorical Variable

		worksheet = workbook.add_worksheet('Categorical Variable Considered')
		row = 1
		col = 0

		worksheet.write(0, 0, 'Categorical Variables considered for feature selection')

		for category in categor_features:
			worksheet.write(row, col, category)
			# worksheet.write(row, col + 1, elm2)
			row += 1

		## Sixth Sheet Additional features


		## Seventh Sheet constraints Settings

		constaints_check = load_jsons['advncsetting']

		worksheet = workbook.add_worksheet('Constraints Settings')
		row = 1
		col = 0
		if len(constaints_check) == 0:
			pass
		else:
			worksheet.write(0, 0, 'Test Store')
			worksheet.write(0, 1, 'State')
			worksheet.write(0, 2, 'Store Type')
			worksheet.write(0, 3, 'Locality')

			for constaints in constaints_check:
				worksheet.write(row, col, constaints['store_sk'])
				worksheet.write(row, col + 1, constaints['state_short'])
				worksheet.write(row, col + 2, constaints['store_type'])
				worksheet.write(row, col + 3, constaints['locality_description'])
				row += 1


		## Eigth Sheet Match results

		controlstring = load_jsons['controlstring']
		control_datas = controlstring['data']['matches']

		matches_data = pd.DataFrame(eval(control_datas))
		# worksheet = workbook.add_worksheet('Match Results')
		workbook.close()

		writer = pd.ExcelWriter(report_name, engine='openpyxl', mode='a')
		writer.book = load_workbook(report_name)  # here is the difference
		matches_data.to_excel(writer, sheet_name='Match Results',index=False)
		writer.close()

		## Nineth Sheet Measurement results

		measureTest = load_jsons1['measureTest']
		measure_data = pd.DataFrame(measureTest)

		writer = pd.ExcelWriter(report_name, engine='openpyxl', mode='a')
		writer.book = load_workbook(report_name)  # here is the difference
		measure_data.to_excel(writer, sheet_name='Measurement Results',index=False)
		writer.close()


		shutil.move(os.path.join('', report_name), os.path.join(config['pathConfig']['trialpath'] ,'trial_'+str(pk) , report_name))

		tempdest = os.path.join('/usr/share/nginx/temp/')
		shutil.copyfile(os.path.join(config['pathConfig']['trialpath'] + '/trial_' + str(pk) + "/", report_name),
						tempdest + report_name)

		results =  {
			"filename" :  report_name
		}

		return json.Response(results,True)


def sampleSize(request):

	import json as j 	
	data = j.loads(request.body)
	datas 		= data['data']
	variable 	= datas['variable']
	effectSize  = datas['effectSize']
	ratio   	= datas['ratio']
	alpha     	= datas['alpha']

	if(variable=='Lift'):		
		variable = 'gross'		
	elif(variable=='Sales'):
		variable = 'sales'
	else:
		variable = 'margin'

	variables = {
		'sales' : 'TotNet',
		'gross' : 'TotNet',
		'margin' : 'TotNet'
	}

	variable = variables[variable]
	powers = np.arange(0.3,1,0.1).round(2)
	effectSizes = np.arange(0.1,1.05,0.05).round(2)
	sampleSizes = pd.DataFrame(index= effectSizes, columns= powers)
	analysis = TTestIndPower().solve_power

	if dataStore[variable].dtype not in [int,float]:
		return 0
    
	data = dataStore[ storeKeys + ['wkEndDate',variable] ]
	data = data[data[variable] > 0]

	data['YoY'] = data.groupby(by= storeKeys)[variable].shift(52)
	data = data.dropna(subset= ['YoY'])
	data['YoY'] = (data[variable] / data['YoY']) - 1
	std = data.groupby(by= storeKeys)['YoY'].median().std()

	effectSize = effectSize/std

	def solveSample(row):
		return pd.Series([analysis(effect_size= row.name, nobs1= None, alpha= alpha, power= power, ratio= ratio) for power in row.index.values]).astype('int')

	sampleSizes[powers]= sampleSizes.apply(solveSample, axis= 1)

	table = sampleSizes.loc[effectSize * std].reset_index().rename(columns={'index' : 'power', effectSize*std : 'sampleSize'})


	finalsampleSizes = sampleSizes.reset_index().to_json(orient='records')
	finaltable = table.reset_index().to_json(orient='records')

	results =  {
			"table" :   finaltable,
			"plot"  : 	finalsampleSizes
		}

	return json.Response(results,True)

def analyseFeatures(request):
	"""
        params:
        ------
            trial           :   trial id unique for a chosen test                               -   str
            teststores      :   ids test stores selected                                        -   list of strings
            metric          :   metric chosen for test. 1 - Wt. avg price,
                                                        2 - % Sales                             -   int 
            startDate       :   start date for feature selection - 'yyyy-mm-dd'                 -   string
            endDate         :   end date for feature selection - 'yyyy-mm-dd'                   -   string
            hierarchy       :   hierarchy level selected. dict containing following keys.       -   python dict

                                        'lvl'       : 1 for Level 1,                -   int
                                                      2 for Level 2,
                                                      3 for Level 3,
                                                      4 for Level 4
                                        'catgories' :  list of categories selected. -   list of strings
                                                      (names as per the file provided)

            itemList        :   ids of items selected                                           
            extraFeatures   :   addtional features uploaded by user                             -   pandas DataFrame (None by default)

                                        'store_sk'  :   ids of stores       -   str (should only contain all stores in the market )
                                        <feature_1> :                       -   float/ str
                                        <feature_2> :                       -   float/ str
                                            :
                                            :
                                        <feature_n> :                       -   float/ str
        
        returns:
        -------
            rSquared        :   overall rSquared of model with selected features                    -   float
            rSquaredAdj     :   rSquared_adj of model with selected features                        -   float
            mape            :   mean absolute percentage error of model with selected features      -   float
            mdape           :   median absolute percentage error of model with selected features    -   float
            features        :   selected features along with following columns                      -   list of strings
            otherFeatures   :   unselected features                                                 -   list of strings

        dependencies:
        ------------
            analyseHelper()   :   filtering
            dataPrep          :   modelData preparation
    """

    # dataStore = analyseHelper(startDate, endDate)

	import json as j 	
	data = j.loads(request.body)
	datas = data['data']
	trial 		= datas['trial']
	testStores  = datas['teststores']
	metric      = datas['metric']	        					
	startDate   = datas['startDate']
	endDate     = datas['endDate']
	hierarchy   = datas['hierarchy']
	itemList    = datas['itemList']
	AdditionalFeatures = datas['extraFeatures']

	if(AdditionalFeatures is not None):
		AdditionalFeatures = pd.DataFrame(eval(AdditionalFeatures))
		AdditionalFeatures= AdditionalFeatures.replace(np.nan, ' ', regex=True)
		extraFeatures = pd.DataFrame.from_dict(AdditionalFeatures, orient='columns')
	else:
		extraFeatures = None

	global dataStore

	try:
		logdf = logs("read", trial=trial)
		logdf.loc[trial, analyseFeatureCols] = [None] * len(analyseFeatureCols)
	except Exception as e:
		return json.Response("Can't able to read log", False)

	try:
		startDate = datetime.strptime(startDate, '%Y-%m-%d') 
		endDate = datetime.strptime(endDate, '%Y-%m-%d') 
		data = dataStore.loc[(dataStore.wkEndDate >= startDate) & (dataStore.wkEndDate <= endDate)]

		# NOTE: setting the store x week as index
		data = data.set_index(storeKeys+['wkEndDate'])
        
        # analyseFeatureCols = [testStores, numTeststores, metric, startDate, endDate, numWeeks, hierarchy, numCategories, numProducts, numProductSelected, extraFeatureShape, extraFeatureFiltered, dataInsufficient, multiCorr, modelData.shape, features, gain, pvalues, rSquared, rSquaredAdj, mape, mdape, analyseFeatureStatus, afAlert]
		logdf.loc[trial,analyseFeatureCols[:8]] = [testStores, len(testStores), metric, startDate, endDate, len(data.reset_index().wkEndDate.unique()), hierarchy, len(hierarchy['categories'])]

		modelData, target, extraFlag, numFeatures, catFeatures, logdf, removeStores  = getFeatures(trial= trial, testStores= testStores, dataStore= data, metric= metric, hierarchy= hierarchy, itemList= itemList, extraFeatures= extraFeatures, logdf= logdf.copy())
		if modelData is None:
			return rSquared, rSquaredAdj, mape, mdape, features, otherFeatures, extraFlag, numFeatures, catFeatures, removeStores
        
		features = pd.DataFrame(columns= ['features', 'gain'])
		features['features'] = modelData.columns.values

		modelData[['target']] = target

		# Scale
		values = StandardScaler().fit_transform(modelData)
		x,y = values[:, :-1],values[:, -1]

		# model
		model = xgb.XGBRegressor()
		model.fit(x, y)

		features['gain'] = model.feature_importances_
		features = features.sort_values(by= ['gain'], ascending= False)

		n_features = int(config['runConfig']['n_features'])

		# retrain the model to see the rsquared, adjRSquared
		x_ = x[:, features.iloc[:n_features].index.values]

		model.fit(x_,y)
		yHat = model.predict(x_)

		absError = np.abs(y - yHat)
		ape = np.divide(absError,y) * 100
		mape = np.mean(ape)
		mdape = np.median(ape)

		yMean = y.mean()
		sst = np.square(y - yMean).sum()
		sse = np.square(y - yHat).sum()
		rSquared = 1- (sse/sst)
		rSquaredAdj = 1 - (((1-rSquared) * (len(y)-1)) / (len(y) - len(x_[0]) - 1)) # 1- ((1-r2) * (n-1) / (n-k-1))

		otherFeatures = features.iloc[n_features:]['features'].sort_index().values
		features = features.iloc[:n_features]

		pValues = None

		if extraFeatures is not None:
			extras = extraFlag.loc[(extraFlag.comments == config['logConfig']['multiCorr'])& (extraFlag.comments == config['logConfig']['corr']),'extraFeatures'].values
			if len(extras) > 0:
				fileName = os.path.join(config['pathConfig']['trialPath'],trial,config['fileConfig']['modelData'])
				dataStore = pd.read_csv(fileName)
				extraFeatures = extraFeatures.set_index(storeKeys)
				dataStore = dataStore.merge(extraFeatures[extras],on= storeKeys)
				dataStore.to_csv(fileName)
				chmod(fileName)
			extraFlag = extraFlag.sort_values(by=['comments'])

			forcedCols = [col for col in modelData.columns if '_trend' in col and col not in extraFeatures.columns and col not in features.features.values]
		else:
			forcedCols = [col for col in modelData.columns if '_trend' in col and col not in features.features.values]
        
		features = pd.DataFrame({features.columns.values[0]: forcedCols},columns= features.columns).append(features)
		otherFeatures = [feature for feature in otherFeatures if feature not in features.features.values]
        
		# analyseFeatureCols = [testStores, numTeststores, metric, startDate, endDate, numWeeks, hierarchy, numCategories, numProducts, numProductSelected, extraFeatureShape, extraFeatureFiltered, dataInsufficient, multiCorr, modelData.shape, features, gain, pvalues, rSquared, rSquaredAdj, mape, mdape, analyseFeatureStatus, afAlert]
		logdf.loc[trial,analyseFeatureCols[-9:]] = [[[col] for col in features.features.values], [[col] for col in features['gain'].values], pValues, rSquared, rSquaredAdj, mape, mdape, success,None]

		logs("save",logdf=logdf)

		selected_stores  = []
		add_feature_list =[]

		finalfeatures = features.reset_index().to_json(orient='records')

		if(extraFlag is not None):
			finalextra = extraFlag.reset_index().to_json(orient='records')
		else:
			finalextra = []

		for add_featus in list(otherFeatures):
			add_feature_list.append(add_featus)

		for select_stores in list(target.index.values):
			selected_stores.append(select_stores)

		results = {
				"rSquared" : rSquared,
				"rSquaredAdj" : rSquaredAdj,
				"teststores" : selected_stores,
				"mape" : mape,
				"mdape" :mdape,
				"features" : finalfeatures,
				"additional_features" : add_feature_list,
				"categor_features": catFeatures,
				"contiue_features" :numFeatures,
				"extraflag" :finalextra,
				"remove_stores": removeStores,
				"status" : "ok"
		}
		return HttpResponse (j.dumps(results))

	except Exception as e:
		_, _, exc_tb = sys.exc_info()
		error = [exc_tb.tb_lineno, os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]]
		logdf.loc[trial, analyseFeatureCols[-2:]] = [e, error]
		logs("save", logdf=logdf.copy())
		return json.Response("Can't able to analyse",False)

	
	
def identifyControls(request):
	"""
	        params:
	        ------
	            trial               :   trial id unique for a chosen test                               -   int
	            teststores          :   ids test stores selected                                        -   list of strings
	            selectedFeatures    :   list of selected features (as per the given list)               -   list of strings
	            excludeControls     :   ids of controls to be excluded, if any uploaded                 -   list of strings
	            n_controls          :   no of contrl stores for each test store                         -   int
	            settings            :   filters based on demographics for control stores selection,     -   pandas DataFrame (None by default)
	                                    with 1 for selceted constrains and np.nan (null) for rest
	                                    having following columns

	                                        'store_sk'                :   column listing testStores' id           -   object (str)
	                                        'state_short'               :   1 for selected stores, null for rest    -   int
	                                        'store_type'                :   1 for selected stores, null for rest    -   int
	                                        'locality_description'      :   1 for selected stores, null for rest    -   int

	            excludeControls     :   list of 'store_sk' of store to be excluded in control store     -   list of strings (empty list by default)

	        returns:
	        -------
	            matches             :   test vs control store mapping with the similarity score         -   pandas DataFrame
	            sales               :   sales correlation table                                         -   pandas DataFrame
	            metric_pvalue       :   p-value for each variable used in calculating test-control      -   pandas DataFrame
	                                    similarity scores

	        dependencies:
	        ------------
	            modelData.csv   :   csv containing all the features for modelling for all stores        - .csv file saved at ['dataPath']
	                                (saved by getFeatures())
	            trialSales.csv  :   csv containing weekly sales for all stores for selected interval    - .csv file saved at ['dataPath']
	                                (saved by getFeatures())
	    """

	import json as j
	data = j.loads(request.body)
	datas 				= data['data']
	trial 				= datas['trial']
	testStores  		= datas['teststores']
	selectedFeatures    = datas['selectedFeatures']
	n_controls   		= datas['n_controls']
	settings_data     	= datas['advance_setting']
	excludeControls   	= datas['excludeControls']

	if(len(settings_data)>0):
		if (settings_data is not None):
			settings_data = pd.DataFrame(settings_data)
			settings = pd.DataFrame.from_dict(settings_data, orient='columns')
		else:
			settings = None
	else:
		settings = None

	try:
		logdf = logs("read", trial=trial)
		logdf.loc[trial, identifyControlCols] = [None] * len(identifyControlCols)
	except:
		return json.Response("Can't able to read log", False)

	# try:
	trial = str(trial)
	selectedFeatures = list(selectedFeatures)
	ex = None
	if len(excludeControls) > 0:
		ex = [[con] for con in excludeControls]
        # ex = excludeControls
    # identifyControlCols = selectedFeatures, storesWithSettings, numStoresWithSettings, nControls, excludeControls, excludeControls.len, identifyControlStatus, icAlert
	logdf.loc[trial,['selectedFeatures', 'nControls', 'excludeControls', 'excludeControls.len']] = np.array([selectedFeatures, n_controls, ex, len(excludeControls)])

	modelData = pd.read_csv(os.path.join(config['pathConfig']['trialPath'],trial, config['fileConfig']['modelData']))
	modelData[storeKeys] = modelData[storeKeys].astype('str')
	modelData = modelData.set_index(storeKeys)
	modelData = modelData[selectedFeatures]

	gower = pd.DataFrame(gf(modelData), index= modelData.index.values, columns= modelData.index.values)
	gower = gower.loc[~gower.index.isin(testStores+excludeControls), testStores]

	global storeChar
    # from scipy.spatial.distance import pdist, squareform
    # eucl = pd.DataFrame(squareform(pdist(normalize(modelData))), index= modelData.index.values, columns= modelData.index.values)
    # eucl = eucl.loc[~eucl.index.isin(testStores+excludeControls), testStores]

	if settings is not None:
		settings = settings[storeKeys+constrainCols].set_index(storeKeys)
		storesWithSettings = settings.dropna(axis=0,how='all').index.values
		logdf.loc[trial,['storesWithSettings','numStoresWithSettings']] = [storesWithSettings, len(storesWithSettings)]

		factors = settings.columns.values
		settings[factors] = np.where(settings[factors] == 1, storeChar.loc[settings.index, factors], settings[factors])
		storeChar_ = storeChar.loc[~storeChar.index.isin(testStores+excludeControls), factors]
        
		def constraints(row):
			rowCols = row.dropna().index.values.tolist()
			storeFactors = storeChar_[rowCols]
			storeFactors[rowCols] = np.where(np.equal(storeFactors[rowCols], row.loc[rowCols]), storeFactors[rowCols], None)
			return storeFactors[rowCols].dropna().index.values

		settings['possibleMatch'] = settings.apply(constraints, axis= 1)
		# settings['controlStores'] = settings.apply(lambda test, eucl= gower, n_controls= n_controls: eucl.loc[test['possibleMatch'], test.name].sort_values().index.values[:n_controls], axis= 1)
		settings['controlStores'] = settings.apply(lambda test, eucl= gower: eucl.loc[test['possibleMatch'], test.name].sort_values().index.values[:20], axis= 1)
		settings = settings.reset_index()
		matches = settings[storeKeys+['controlStores']]
	else:
		matches = pd.DataFrame(testStores, columns= storeKeys)
		# matches['controlStores'] = matches.apply(lambda test, eucl= gower, n_controls= n_controls: eucl[test[storeKeys[0]]].sort_values().index.values[:n_controls], axis= 1)
		matches['controlStores'] = matches.apply(lambda test, eucl= gower: eucl[test[storeKeys[0]]].sort_values().index.values[:20], axis= 1)

	matches = matches.explode('controlStores')
	matches = matches.reset_index(drop= True)
	matches['similarity'] = matches.apply(lambda pair, dist= gower: 1-dist.loc[pair['controlStores'], pair[storeKeys[0]]], axis= 1)
    
	trialSales = pd.read_csv(os.path.join(config['pathConfig']['trialPath'], trial, config['fileConfig']['trialSales']))
	trialSales[storeKeys] = trialSales[storeKeys].astype('str')
	trialSales = trialSales.set_index(['wkEndDate'])

	def salesCorr(row):
		test = trialSales.loc[trialSales[storeKeys[0]] == row[storeKeys[0]], 'TotNet'].reset_index()
		control = trialSales.loc[trialSales[storeKeys[0]] == row['controlStores'], 'TotNet'].reset_index()
		return test['TotNet'].corr(control['TotNet'])

	matches['salesCorrelation'] = matches.apply(salesCorr, axis= 1)
		# matches = matches[~((matches['salesCorrelation'] < 0) & (matches['nextPositive'] < 0.1))]
		# matches = matches[matches['salesCorrelation'] > 0] 
	matches['rank'] = matches.groupby(storeKeys).cumcount() +1
	matches = matches.loc[matches['rank'] <= n_controls]

	allStores = pd.DataFrame(list(matches[storeKeys[0]].unique()) + list(matches.controlStores.unique()), columns= storeKeys)
	allStores = allStores.merge(trialSales.groupby(by= storeKeys).mean(),on= storeKeys).rename(columns={'TotNet' : 'storeSales'})
	allStores = allStores.set_index(storeKeys)

	trialSales['TotNet'] = trialSales.groupby(by= storeKeys).transform(lambda sales: (sales - sales.mean())/sales.std())

	trialSales = trialSales.reset_index()
	trialSales = trialSales.pivot_table(index=['wkEndDate'],columns=storeKeys)['TotNet']
	trialSales = trialSales.round(2)
	columns = ['testStores','sales']+reduce(lambda a,b:a+b,[['control_'+str(i+1),'controlSales_'+str(i+1)] for i in range(n_controls)])
	storeCols = ['testStores']+[col for col in columns if col.startswith('control_')]
	salesCols = ['sales']+[col for col in columns if col.startswith('controlSales_')]
	sales = pd.DataFrame(columns= columns)
	sales.index.name = 'wkEndDate'
	tempdf = pd.DataFrame(index= trialSales.index,columns=columns)
	for testStore in testStores:
		stores = [testStore] + matches.loc[matches[storeKeys[0]] == testStore, 'controlStores'].values.tolist()
		tempdf.loc[:, storeCols] = (storeChar.loc[stores, 'storeName'].astype('str').index + '_' + storeChar.loc[stores, 'storeName'].astype('str')).values
		# tempdf.loc[:,storeCols] = stores
		tempdf.loc[:, salesCols] = trialSales[stores].values
		sales = sales.append(tempdf)
    
	# filtering the modelData for the required test and control stores and the selected features - with the test & control mapping
	testControl = matches.copy()
	testControl = pd.melt(testControl, id_vars = ['rank'])
	testControl.columns = ['rank','testControl']+ storeKeys
	testControl = pd.merge(testControl,modelData[selectedFeatures].reset_index(), how= 'left',on= storeKeys)
	# transpose to get the metrics in a single column
	testControl = testControl.set_index(storeKeys + ['testControl','rank']).stack().reset_index()
	testControl.columns = storeKeys + ['testControl','rank','metric','metric_value']

	# calculating the p-value for each metric
	dict1={}
	for j in testControl['rank'].unique().tolist() :
		dict2={}
		for i in selectedFeatures :
			_, p = ttest_ind(testControl[(testControl['testControl'] == 'controlStores') & (testControl['metric'] == i) & (testControl['rank'] == j)]['metric_value'],testControl[(testControl['testControl'] == storeKeys[0]) & (testControl['metric'] == i)]['metric_value'], equal_var = False)
			dict2[i]=p
		dict1[j]=dict2
	metric_pvalue = pd.DataFrame.from_dict(dict1).reset_index().rename(columns={"index":"features"})
	metric_pvalue.columns = ['features']+['control_'+str(i+1) for i in range(n_controls)]

	matches = matches.rename(columns= {storeKeys[0]: 'testStores'})
	matches = matches.round(2)

	tests = reduce(lambda df1,df2 : pd.merge(df1,df2,on= storeKeys),[storeChar.loc[testStores,constrainCols],modelData.loc[testStores,selectedFeatures],allStores]).reset_index().rename(columns={storeKeys[0]:'testStores'})
	controls = reduce(lambda df1,df2 : pd.merge(df1,df2,on= storeKeys),[storeChar.loc[matches.controlStores.unique(),constrainCols], modelData.loc[matches.controlStores.unique(),selectedFeatures], allStores]).reset_index().rename(columns={storeKeys[0]:'controlStores'})
	tests.columns= ['testStores']+[col+'_test' for col in tests.columns if col!='testStores']
	controls.columns= ['controlStores']+[col+'_control' for col in controls.columns if col!='controlStores']

	order = tests.columns.tolist() + controls.columns.tolist() + [col for col in matches.columns if col not in ['testStores','controlStores']]
	matches = matches.merge(tests,on=['testStores']).merge(controls,on=['controlStores'])[order]

	fileName = os.path.join(config['pathConfig']['trialPath'], trial, config['fileConfig']['matches'])
	matches.to_csv(fileName, index=False)
	chmod(fileName)

	nUniqueControls = matches.controlStores.unique().shape[0]
	meanSimilarity = matches.similarity.mean()

	matches['testStores'] = (storeChar.loc[matches['testStores'], 'storeName'].astype('str').index + '_' + storeChar.loc[matches['testStores'], 'storeName'].astype('str')).values
	matches['controlStores'] = (storeChar.loc[matches['controlStores'], 'storeName'].astype('str').index + '_' + storeChar.loc[matches['controlStores'], 'storeName'].astype('str')).values
	sales = sales.reset_index()
	metric_pvalue = metric_pvalue.round(2)

	logdf.loc[trial,identifyControlCols[-2:]] = [success, None]
	logs("save",logdf=logdf)

	finalmatches = matches.round(2).to_json(orient='records')
	finalsales = sales.reset_index().to_json(orient='records')
	# finalmatchresults = matchResults.to_json(orient='records')
	# finalcolumns = columns.to_json(orient='records')
	results = {
	"matches": finalmatches,
	# "reports": finalmatchresults,
	"columns": matches.columns.tolist(),
	"sales": finalsales,
	"metric": metric_pvalue.round(2).to_dict(),
	"UniqueControls" :nUniqueControls,
	"meanSimilarity" :meanSimilarity
	}

	return json.Response(results)
	# except Exception as e:
	# 	_, _, exc_tb = sys.exc_info()
	# 	error = [exc_tb.tb_lineno, os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]]
	# 	logdf.loc[trial, identifyControlCols[-2:]] = [e, error]
	# 	logs("save", logdf=logdf.copy())
	# 	return json.Response("Can't able to analyse", False)



def liftAnalysis(request):
# def liftAnalysis(trial, preTest, postTest, metrics):
	"""
	        params:
	        ------
	            trial       :   trial id unique for a chosen test                                       -   int
	            storeMap    :   test and control stores mapping with follwing columns                   -   pandas DataFrame
	                                'testStores'     :  testStores (repeats for all controls)
	                                'controlStores'  :  controlStores
	                                'rank'			 :  rank of controlStores
	            preTest     :   dict containing following keys                                          -   python dict
	                                'startDate' : start date of pretest   -   string in 'yyyy-mm-dd' format
	                                'endDate'   : end date of pretest     -   string in 'yyyy-mm-dd' format
	            postTest    :   dict containing following keys                                          -   python dict
	                                'startDate' : start date of posttest  -   string in 'yyyy-mm-dd' format
	                                'endDate'   : end date of posttest    -   string in 'yyyy-mm-dd' format
	            metrics     :   list of metrics selected (the mapping should be as follows)             -   list of int
	                                    1   :   transaction count
	                                    2   :   gross sales
	                                    3   :   net sales
	                                    4   :   food and paper cost
	                                    5   :   gp percent
	                                    6   :   unit count

	        returns:
	        -------
	            meaureTest   :  Dataframe with lift calculated using aggregate values for each metric
	                            during the pre and post test period for every test and control store.   - pandas DataFrame
	            metric_pvalue:  p-value for every metric                                                - float
	        dependencies:
	        ------------
	            dataStore.csv   :   file containing store x week data located at the path           -   .csv file
	                                mentioned in config.ini as ['dataPath']
	    """

	## Preparing the control-test data for the selected trial
	# pull saved trial data table for the given trial - store grp, store_sk, storename, rank, similarity value, sales correlation
	# TODO: change colummn names based on what output comes

	import json as j
	data = j.loads(request.body)
	datas 				= data['data']
	trial 				= datas['trial']
	metrics  			= datas['metrics']
	preTest    			= datas['preTest']
	postTest   			= datas['postTest']
	storeMap			= datas['storeMap']


	try:
		logdf = logs("read", trial=trial)
		logdf.loc[trial, liftAnalysisCols] = [None] * len(liftAnalysisCols)
	except:
		return json.Response("Can't able to read log", False)


	# try:
	trial = str(trial)
    # liftAnalysisCols = prePeriod, postPeriod, metrics, liftAnalysisStatus, laAlert
    
	logdf.loc[trial,liftAnalysisCols[:2]] = [preTest, postTest]
	if storeMap is None:
		storeMap = pd.read_csv(os.path.join(config['pathConfig']['trialPath'], trial, config['fileConfig']['matches']))
		storeMap['testStores'] = storeMap['testStores'].astype('str')
		storeMap['controlStores'] = storeMap['controlStores'].astype('str')           
	else:
		storeMap = pd.DataFrame(eval(storeMap))
		storeMap = pd.DataFrame.from_dict(storeMap, orient='columns')
		storeMap['testStores'] = storeMap['Test store ID (storeID)'].astype('str')
		storeMap['controlStores'] = storeMap['Control store ID (storeID)'].astype('str')
		storeMap['rank'] = storeMap['Rank']

	storeMap = storeMap[['testStores','controlStores','rank']]
	storeMap['testStores'] = storeMap['testStores'].astype('str')
	storeMap['controlStores'] = storeMap['controlStores'].astype('str')
	trialTest = storeMap[['testStores']].drop_duplicates().reset_index(drop= True)
	trialTest['storeGroup'] = 'TEST'
	trialControl = storeMap[['testStores', 'controlStores', 'rank']].reset_index(drop= True)
	trialControl['storeGroup'] = 'CONTROL'

	# append the test and control tables 
	trialData = pd.concat([trialTest, trialControl], axis= 0, sort= True).sort_values(by= 'testStores')
	# replacing the nans in the mapping columns for test stores with test id
	trialData[storeKeys[0]] = trialData.apply(lambda x : x['testStores'] if x['storeGroup'] == 'TEST' else x['controlStores'], axis= 1)
	trialData = trialData.drop(columns=['controlStores'])
    
    # mapping the metric to the actual name in the table 
	metricsMap = {
        1 : 'TotNet',
        2 : 'gross_sales',
        3 : 'unit_count_adj',
        4 : 'total_txns',
        5 : 'gp_percent'
    }

	metricsName = {
        'total_txns'        : 'Transaction_Count',
        'gross_sales'       : 'Gross_Sales',
        'TotNet'            : 'Net_Sales',
        'gp_percent'        : 'Gp_Percent',
        'unit_count_adj'    : 'Unit_Count'
    }
    
	metrics = [metricsMap[metric] for metric in metrics]
	logdf.loc[trial,liftAnalysisCols[2]] = [metrics]

	# Formatting the input dates
	preStartDate = datetime.strptime(preTest['startDate'], '%Y-%m-%d') 
	preEndDate = datetime.strptime(preTest['endDate'], '%Y-%m-%d') 
	postStartDate = datetime.strptime(postTest['startDate'], '%Y-%m-%d') 
	postEndDate = datetime.strptime(postTest['endDate'], '%Y-%m-%d') 

	# Fetching data from dataStore - read the pre_test and post_test start to end weeks data from the dataStore
	global dataStore

	# Getting the list of  all the test and control stores - to pull from dataStore
	testStores = storeMap['testStores'].unique().tolist()
	controlStores = storeMap['controlStores'].unique().tolist()
	# from the dataStore - pull the corresponding test and control store ids and filter the datastore for these ids & metric colums selected
	measureTest = dataStore[dataStore[storeKeys[0]].isin(testStores + controlStores)][[storeKeys[0], 'wkEndDate'] + metrics ].reset_index(drop= True)
	for metric in metrics:
		measureTest['YoY_'+metric] = measureTest.groupby(by=storeKeys)[metric].shift(52)
		measureTest[metric] = (measureTest[metric] / measureTest['YoY_'+metric]) - 1

    # filter for the required time period
	measureTest = measureTest[((measureTest['wkEndDate'] >= preStartDate) & (measureTest['wkEndDate'] <= preEndDate)) | ((measureTest['wkEndDate'] >= postStartDate) & (measureTest['wkEndDate'] <= postEndDate))].reset_index(drop= True)
    
    # label by period into pre and post window
	measureTest['prepost'] = np.where(((measureTest['wkEndDate'] >= preStartDate) & (measureTest['wkEndDate'] <= preEndDate)), 'pre_period', 'post_period')
    
    # aggregate metrics to store level pre and post period  - taking average
    # measureTest = measureTest.groupby([storeKeys[0], 'prepost']).agg('mean')[metrics].reset_index()
	measureTest = measureTest.groupby([storeKeys[0], 'prepost']).agg('mean').reset_index()

    # pdb.set_trace()
	for metric in metrics:
		measureTest[metric] = (measureTest[metric] - measureTest['YoY_'+metric])
    
	# transpose and get pre and post into columns and metric into rows
	measureTest = measureTest.set_index([storeKeys[0], 'prepost']).stack().reset_index()
	measureTest.columns = [storeKeys[0], 'prepost', 'metric', 'metric_value']
	measureTest = measureTest.replace(metricsName)
	measureTest = pd.pivot_table(measureTest, index= [storeKeys[0], 'metric'], columns= 'prepost', values= 'metric_value').reset_index()
	measureTest = pd.merge(trialData, measureTest, on= storeKeys, how= 'left')

	# standardising the metrics by the range of the test pre and post periods
	# measureTest['pre_period'] =  measureTest['pre_period']/ round(((preEndDate - preStartDate).days)/7)
	# measureTest['post_period'] =  measureTest['post_period']/ round(((postEndDate - postStartDate).days)/7)

	global storeChar
    # calcualte the lift
	measureTest['lift'] = (measureTest['post_period']/measureTest['pre_period']-1)*100
	delta = measureTest.loc[measureTest['storeGroup'] == 'CONTROL'].groupby(["testStores","metric"])["lift"].mean().reset_index()
	delta = pd.merge(delta, measureTest.loc[measureTest['storeGroup'] == 'TEST',['testStores','metric','lift']],on= ['testStores','metric'])
	delta['delta'] = delta['lift_y'] - delta['lift_x']
	delta.drop(['lift_x','lift_y'],axis=1,inplace=True)
	delta['testStores'] = (storeChar.loc[delta['testStores'], 'storeName'].astype('str').index + '_' + storeChar.loc[delta['testStores'], 'storeName'].astype('str')).values
	delta = delta.sort_values(by=['delta'])
	metric_pvalue = pd.DataFrame(columns = ['measureMetric','pValue','delta'])
	metric_pvalue['measureMetric'] = metrics
	metric_pvalue = metric_pvalue.replace(metricsName)
	# calcualte the p-value 
	measureTest = measureTest.dropna(subset=['lift'])
	for i in measureTest.metric.unique() :
		_, p = ttest_ind(measureTest[(measureTest['storeGroup'] == 'CONTROL') & (measureTest['metric'] == i)]['lift'], measureTest[(measureTest['storeGroup'] == 'TEST') & (measureTest['metric'] == i)]['lift'], equal_var = False)
		diff = delta.groupby("metric")["delta"].mean()[i]
		delta.loc[len(delta)] = ['average',i,diff]
		metric_pvalue.loc[metric_pvalue["measureMetric"]==i,["pValue","delta"]] = [p, diff] 

	measureTest['storeName'] = (storeChar.loc[measureTest[storeKeys[0]], 'storeName'].astype('str')).values
	measureTest = measureTest.round(2)
	# metric_pvalue = metric_pvalue.round(4)
	delta = delta.round(4)

	# liftAnalysisCols = prePeriod, postPeriod, metrics, pValue, delta, liftAnalysisStatus, laAlert

	logdf.loc[trial,liftAnalysisCols[-4:]] = [metric_pvalue[metric_pvalue.columns.values[1]].values, metric_pvalue[metric_pvalue.columns.values[-1]].values, success,None]

	logs("save",logdf=logdf)


	variable =[]
	metric = []
	deltas =[]

	import math
	for	i in metric_pvalue.measureMetric:
		variable.append(i)
	for j  in metric_pvalue.pValue:
		if(math.isnan(j)):
			metric.append("-0")
		else:
			metric.append(round(j,2))
	for k  in metric_pvalue.delta:
		deltas.append(round(k,1))

	finalmeasureTest = measureTest.round(2).reset_index().to_json(orient='records')
	finaldelta = delta.round(4).reset_index().to_json(orient='records')

	results = {
		"measureTest": finalmeasureTest,
		"variable" : variable,
		"metric": metric,
		"delta" : deltas,
		"deltagraph" :finaldelta

	}

	return json.Response(results)

	# except Exception as e:
	# 	_, _, exc_tb = sys.exc_info()
	# 	error = [exc_tb.tb_lineno, os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]]
	# 	logdf.loc[trial, liftAnalysisCols[-2:]] = [e, error]
	# 	logs("save", logdf=logdf.copy())
	# 	return json.Response("Can't able to analyse", False)

